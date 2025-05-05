#!/usr/bin/env python3
#
# updated by ...: Loreto Notarantonio
# Date .........: 05-05-2025 17.49.29
#


import sys; sys.dont_write_bytecode = True
import os
from pathlib import Path
from benedict import benedict
from types import SimpleNamespace
from enum import Enum
import pandas as pd

import  openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pprint import pprint as pp

self=sys.modules[__name__]

import ln_pandasExcel_Class as lnExcel
import lnUtils
import dictUtils
from ln_pandasExcel_Class import workBbookClass, sheetClass
import commonFunctions


def setup(gVars: (dict, SimpleNamespace)):
    global gv
    gv=gVars
    gv.logger.caller(__name__)






#################################################################
#
#################################################################
# def create(d: dict, level: int, sh_name: str):
def create(d: dict, hierarchy_level):
    level = hierarchy_level.value
    sh_name = hierarchy_level.name

    ## catturiamo tutti i records fino al livello di agent creando dei keypath
    keypaths = dictUtils.chunckList(gv.keypaths_list, item_nrs=level)

    ### --- remove duplicate entries
    hier_keypaths = lnUtils.removeListOfListDuplicates(list_of_lists=keypaths, keep_order=True)


    ### --- create title row
    title_row = commonFunctions.prepareTitleRow(index=level)


    # -------------------------------------------------------------------------------------
    # --- creazione delle colonne da partenr in poi (con i dati dei contratti)
    #--- creiamo un dictionary con key=partner
    # --- {
    # ---   "edison": [0,0,0,0,0], somma dei vaori dei singoli agenti
    # ---   "....":   [0,0,0,0,0]
    # ---   }
    # --- con questi dati andrò a creare delle righe sotto il Team Manager
    # -------------------------------------------------------------------------------------
    sheet_rows = [] # righe del foglio excel
    row_to_be_colored = [] # da evidenziare
    target_level = gv.HIERARCHY.Agente.value - (level+1)

    for index in range(len(hier_keypaths)):
        partner_col_data = gv.myDict()
        # tm_somma=gv.default_result_cols[:] ### conterrà la somma dei vari partner
        tm_somma=commonFunctions.result_columns() ### conterrà la somma dei vari partner

        this_level_kp=hier_keypaths[index] ## riga corrente

        manager_name=this_level_kp[-1]
        gv.logger.info("analysing data for %s %s:", sh_name, manager_name)

        if d_data := d[this_level_kp]:
            if target_level < 0: ### siamo a livello di Agente. non dobbiamo raggrupparli
                commonFunctions.partnerData(agent_data=d_data, partner_column=partner_col_data, somma=tm_somma)
            else:
                agent_list = dictUtils.get_keys_at_level(d_data, target_level=target_level, current_level=0)
                gv.logger.info("analysing data for Agents: %s:", agent_list)

                ### --- aggiunge tutte le rige dei  partner sommati per i relativi agenti
                commonFunctions.processAgentList(agent_list=agent_list, partner_column=partner_col_data, somma=tm_somma)

            new_row = this_level_kp[:]

            ### --- riga con i totali per teamManager
            new_row.append('Summary')
            new_row.extend(tm_somma)
            sheet_rows.append(new_row)
            row_to_be_colored.append(len(sheet_rows)+1) ### aggiungere il titolo

            if partner_col_data:
                for partner, data in partner_col_data.items():
                    gv.logger.notify("    agent data has been found")
                    new_row = this_level_kp[:]
                    new_row.append(partner)
                    new_row.extend(data)
                    sheet_rows.append(new_row)
            else:
                gv.logger.warning("    NO agent data found")

    # --- @Loreto:  eliminiamo le celle che hanno valore == cella superire
    rows_data = dictUtils.compact_list(data=sheet_rows, max_items=level, replace_str='-')

    ### - creiamo il dataFrame
    df = pd.DataFrame(
            # columns = colonne_gerarchia[:inx+1],
            columns = title_row,
            data    = rows_data
        )

    excel_file_path=gv.args.output_agenti_filename
    excel_file_path=gv.excel_filename
    # lnExcel.addSheet(filename=gv.args.output_agenti_filename, sheets=[sh_name], dataFrames=[df], sheet_exists="replace", mode='a')
    lnExcel.addSheet(filename=excel_file_path, sheets=[sh_name], dataFrames=[df], sheet_exists="replace", mode='a')


    ### --- crea il range del Manager e relativi risultati
    cell_range=[]
    for col in range(level, len(title_row)+1):
        row_cells = [(row, col) for row in row_to_be_colored]
        cell_range.extend(row_cells)


    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb[sh_name]

    commonFunctions.setColumnSize(ws)
    commonFunctions.setTitle(ws)
    commonFunctions.setCellsColor(ws, cells=cell_range, color='ffffa6')

    ws.freeze_panes = ws['B2'] ## Freeze everything to left of B (that is A) and no columns to feeze
    wb.save(excel_file_path)




def agentiNonTrovati(agents: list):
    ### - creiamo il dataFrame

    sh_name = "Agents NOT found"
    title_row = ["Agenti NON trovati"]
    rows_data = []
    for agent_name in list(agents.keys()):
        rows_data.append(agent_name)

    df = pd.DataFrame(
            # columns = colonne_gerarchia[:inx+1],
            columns = title_row,
            data    = rows_data
        )

    excel_file_path=gv.args.output_agenti_filename
    excel_file_path=gv.excel_filename
    # lnExcel.addSheet(filename=gv.args.output_agenti_filename, sheets=[sh_name], dataFrames=[df], sheet_exists="replace", mode='a')
    lnExcel.addSheet(filename=excel_file_path, sheets=[sh_name], dataFrames=[df], sheet_exists="replace", mode='a')

    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb[sh_name]
    commonFunctions.setColumnSize(ws)
    ws.freeze_panes = ws['B2'] ## Freeze everything to left of B (that is A) and no columns to feeze
    wb.save(excel_file_path)