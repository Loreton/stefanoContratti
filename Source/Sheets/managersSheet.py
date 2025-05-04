#!/usr/bin/env python3
#
# updated by ...: Loreto Notarantonio
# Date .........: 04-05-2025 20.07.53
#


import sys; sys.dont_write_bytecode = True
import os
from pathlib import Path
from benedict import benedict
from types import SimpleNamespace
from enum import Enum
import pandas as pd

import  openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pprint import pprint as pp

self=sys.modules[__name__]

import ln_pandasExcel_Class as lnExcel
import lnUtils
import dictUtils
from ln_pandasExcel_Class import workBbookClass, sheetClass
import commonFunctions


def setup(gVars: (dict, SimpleNamespace)):
    global gv
    gv=gVars
    gv.logger.caller(__name__)






#################################################################
#
#################################################################
def createSheet(d: dict, level: int, sh_name: str):

    # separator = '#'

    ## catturiamo tutti i records fino al livello di agent creando dei keypath
    keypaths = dictUtils.chunckList(gv.keypaths_list, item_nrs=level)

    ### --- remove duplicate entries
    managers_keypaths = lnUtils.removeListOfListDuplicates(list_of_lists=keypaths, keep_order=True)
    # for item in managers_keypaths: print(item)


    title_row = commonFunctions.prepareTitleRow(index=level)


    # -------------------------------------------------------------------------------------
    # --- @Loreto: riempiamo le colonne dati con il valori agente
    #--- creiamo un dictionary con key=partner
    # --- {
    # ---   "edison": [0,0,0,0,0], somma dei vaori dei singoli agenti
    # ---   "....":   [0,0,0,0,0]
    # ---   }
    # --- con questi dati andrò a creare delle righe sotto il Team Manager
    # -------------------------------------------------------------------------------------
    sheet_rows = [] # righe del foglio excel
    row_to_be_colored = []

    for index in range(len(managers_keypaths)):
        partner_col_data = gv.myDict()
        tm_somma=gv.default_result_cols[1:] ### conterrà la somma dei vari partner
        manager_kp=managers_keypaths[index] ## riga corrente

        manager_name=manager_kp[-1]
        gv.logger.info("analysing data for Manager %s:", manager_name)

        agent_list = dictUtils.get_keys_at_level(d[manager_kp], target_level=1, current_level=0)
        gv.logger.info("analysing data for Agents: %s:", agent_list)

        ### --- aggiunge tutte le rige dei  partner sommati per i relativi agenti
        commonFunctions.processAgentList(agent_list=agent_list, partner_column=partner_col_data, somma=tm_somma)

        new_row = manager_kp[:]

        ### --- riga con i totali per teamManager
        new_row.append('somma')
        new_row.extend(tm_somma)
        sheet_rows.append(new_row)
        row_to_be_colored.append(len(sheet_rows)+1) ### aggiungere il titolo

        if partner_col_data:
            for partner, data in partner_col_data.items():
                gv.logger.notify("    agent data has been found")
                new_row = manager_kp[:]
                new_row.append(partner)
                new_row.extend(data)
                sheet_rows.append(new_row)
        else:
            gv.logger.warning("    NO agent data found")


    # --- @Loreto:  eliminiamo le celle che hanno valore == cella superire
    rows_data = dictUtils.compact_list(data=sheet_rows, max_items=level, replace_str='-')

    ### - creiamo il dataFrame
    df = pd.DataFrame(
            # columns = colonne_gerarchia[:inx+1],
            columns = title_row,
            data    = rows_data
        )


    lnExcel.addSheet(filename=gv.args.output_agenti_filename, sheets=[sh_name], dataFrames=[df], sheet_exists="replace", mode='a')


    ### --- crea il range del Manager e relativi risultati
    cell_range=[]
    for col in range(gv.COLS.Manager.value, len(title_row)+1):
        row_cells = [(row, col) for row in row_to_be_colored]
        cell_range.extend(row_cells)

    file_path = gv.args.output_agenti_filename
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sh_name]

    commonFunctions.setColumnSize(ws)
    commonFunctions.setTitle(ws)
    commonFunctions.setCellsColor(ws, cells=cell_range, color='ffffa6')

    ws.freeze_panes = ws['B2'] ## Freeze everything to left of B (that is A) and no columns to feeze
    wb.save(file_path)

