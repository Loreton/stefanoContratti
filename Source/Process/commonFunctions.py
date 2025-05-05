#!/usr/bin/env python3
#
# updated by ...: Loreto Notarantonio
# Date .........: 05-05-2025 17.44.54
#


import sys; sys.dont_write_bytecode = True
import os
from pathlib import Path
from benedict import benedict
from types import SimpleNamespace
from enum import Enum
import pandas as pd

import  openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --- @Loreto: my lib
import ln_pandasExcel_Class as lnExcel
import lnUtils
import dictUtils
from ln_pandasExcel_Class import workBbookClass, sheetClass




def setup(gVars: (dict, SimpleNamespace)):
    global gv
    gv=gVars
    gv.logger.caller(__name__)





########### Excel function ----
def prepareTitleRow(index: int):
    # --- @Loreto: prepariamo il titolo
    title_row = gv.colonne_gerarchia[:index]
    title_row.append("Partner")
    # for col_name in gv.colonne_dati:
    for col in gv.dataCols:
        title_row.append(col.name)
    return title_row



def result_columns():
    # --- aggiungiamo le colonne contenenti i risultati di default (=0)
    default_result_cols = []
    # for col_name in gv.colonne_dati:
    for col in gv.dataCols:
        default_result_cols.append(0) ### - Valore di default
    # default_result_cols[0] = "" ### replace with blank value
    return default_result_cols



def setTitle(ws):
    gray = 'b2b2b2'
    # formatting the header columns, filling red color
    for col in range(1, ws.max_column + 1):
       cell_header = ws.cell(1, col)
       cell_header.fill = PatternFill(start_color=gray, end_color=gray, fill_type="solid") #used hex code for red color


# Auto-adjust Excel column widths
def setColumnSize(ws):

    from openpyxl.utils import get_column_letter
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].auto_size = True
    return

    # for col in ws.columns:
    #     max_length = 0
    #     column = col[0].column_letter  # Get the column name (e.g., 'A')
    #     for cell in col:
    #         try:
    #             if cell.value:
    #                 max_length = max(max_length, len(str(cell.value)))
    #         except:
    #             pass
    #     adjusted_width = (max_length + 2)
    #     ws.column_dimensions[column].width = adjusted_width
    #     gv.logger.notify("setting %s col_width to: %s", column, adjusted_width)



############################################################
# cell_range = [ (row1, col1), (row2, col2), ...]
############################################################
def setCellsColor(ws, cells: list, color='ffffa6'):
    # light_yellow_3 = 'ffffa6'
    # my_color = light_yellow_3

    for row, col in cells:
        curr_cell = ws.cell(row, col)
        curr_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid") #used hex code for red color






def partnerData(agent_data: dict, partner_column: dict, somma: list):
    dc = gv.dataCols
    for partner, data in agent_data.items():
        if not partner in partner_column:
            partner_column[partner] = gv.default_result_cols[:] ## skip partner name
            # partner_col_data[partner][0] = partner

        ptr = partner_column[partner]
        # import pdb; pdb.set_trace() # by Loreto
        # ptr[0] += data["processati"]
        # ptr[1] += data["discarded"]
        # ptr[2] += data["excluded"]
        # ptr[3] += data["totale"]
        # ptr[4] += data["confermato"]
        # ptr[5] += data["attivazione"]
        # ptr[6] += data["back"]
        # ptr[7] += data["rid"]

        ptr[0] += data[dc.PROCESSATI.name]
        ptr[1] += data[dc.EXCLUDED.name]
        ptr[2] += data[dc.INSERITI.name]
        ptr[3] += data[dc.SCARTATI.name]
        ptr[4] += data[dc.TOTALE.name]
        ptr[5] += data[dc.CONFERMATI.name]
        ptr[6] += data[dc.ATTIVAZIONE.name]
        ptr[7] += data[dc.BACK.name]
        ptr[8] += data[dc.RID.name]
        ptr[9] += data[dc.VAS.name]
        ptr[10] += data[dc.SIM.name]
        ptr[11] += data[dc.TV.name]



        ### --- aggiorniamo il totale
        somma[0] += data[dc.PROCESSATI.name]
        somma[1] += data[dc.EXCLUDED.name]
        somma[2] += data[dc.INSERITI.name]
        somma[3] += data[dc.SCARTATI.name]
        somma[4] += data[dc.TOTALE.name]
        somma[5] += data[dc.CONFERMATI.name]
        somma[6] += data[dc.ATTIVAZIONE.name]
        somma[7] += data[dc.BACK.name]
        somma[8] += data[dc.RID.name]
        somma[9] += data[dc.VAS.name]
        somma[10] += data[dc.SIM.name]
        somma[11] += data[dc.TV.name]




        # somma[0] += data["processati"]
        # somma[1] += data["discarded"]
        # somma[2] += data["excluded"]
        # somma[3] += data["totale"]
        # somma[4] += data["confermato"]
        # somma[5] += data["attivazione"]
        # somma[6] += data["back"]
        # somma[7] += data["rid"]







def processAgentList(agent_list: list, partner_column: dict, somma: list):
    for agent_name in agent_list:
        gv.logger.info("    agent: %s", agent_name)
        if agent_data:=gv.agent_results.get(agent_name): # se presente....
            partnerData(agent_data=agent_data, partner_column=partner_column, somma=somma)

